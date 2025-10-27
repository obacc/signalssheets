#!/usr/bin/env python3
"""
SignalsSheets Excel Generator
Creates FREE and BASICO tier Excel files with Trinity Method trading signals
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os

# Brand Kit Colors (ARGB format for openpyxl)
COLORS = {
    'primary': 'FF1E3A8A',      # Azul Índigo (30, 58, 138)
    'success': 'FF10B981',      # Verde Señal (16, 185, 129)
    'danger': 'FFEF4444',       # Rojo Alerta (239, 68, 68)
    'warning': 'FFF59E0B',      # Amarillo (245, 158, 11)
    'text_primary': 'FF1E293B', # Texto principal
    'text_secondary': 'FF64748B', # Texto secundario
    'bg_light': 'FFF8FAFC',     # Background claro
    'border': 'FFE2E8F0',       # Borders
    'card_bg': 'FFEFF6FF',      # Card background (light blue)
    'upgrade_bg': 'FFFED7AA'    # Upgrade banner (light orange)
}

# VBA Code for both files
VBA_CODE = """Option Explicit

Sub RefreshAllQueries()
    Application.ScreenUpdating = False
    Application.StatusBar = "Actualizando datos de Cloudflare..."

    On Error Resume Next
    ActiveWorkbook.Queries.Refresh
    DoEvents

    Application.StatusBar = "¡Datos actualizados! " & Now()
    Application.Wait (Now + TimeValue("0:00:02"))
    Application.StatusBar = False

    Application.ScreenUpdating = True

    MsgBox "Datos actualizados correctamente ✓", vbInformation, "SignalsSheets"
End Sub

Sub MakeVeryHidden()
    ' Make data sheets very hidden (not accessible via UI)
    On Error Resume Next
    Sheets("Config").Visible = xlSheetVeryHidden
    Sheets("DATA_Signals").Visible = xlSheetVeryHidden
    Sheets("DATA_Prices").Visible = xlSheetVeryHidden
    Sheets("DATA_Context").Visible = xlSheetVeryHidden
End Sub
"""

def create_config_sheet(wb, tier="FREE"):
    """Create Config sheet (very hidden)"""
    ws = wb.create_sheet("Config", 0)

    # Data
    token = "tok_free_placeholder_123" if tier == "FREE" else "tok_basico_placeholder_456"
    config_data = [
        ["TOKEN", token],
        ["API_BASE", "https://api.signalssheets.com"],
        ["TIER", tier],
        ["EMAIL", "user@example.com"],
        ["VERSION", "1.0.0"],
        ["LAST_UPDATE", datetime.now()]
    ]

    for row in config_data:
        ws.append(row)

    # Format
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40

    # Will be set to very hidden after creation
    ws.sheet_state = 'hidden'

    return ws

def create_data_signals_sheet(wb):
    """Create DATA_Signals sheet (very hidden)"""
    ws = wb.create_sheet("DATA_Signals")

    # Headers
    headers = ["ticker", "company", "sector", "signal", "trinity_score",
               "lynch_score", "oneil_score", "graham_score", "price", "target",
               "stop_loss", "tp1", "tp2", "potential_return", "risk_reward",
               "author_dominant", "updated_at"]
    ws.append(headers)

    # Dummy data
    dummy_data = [
        ["NVDA", "NVIDIA", "Tech", "BUY", 95, 92, 98, 95, 485.00, 520.00, 475.00, 510.00, 530.00, 0.072, 2.2, "O'Neil", datetime.now()],
        ["AAPL", "Apple", "Tech", "BUY", 92, 88, 95, 90, 175.00, 185.00, 171.00, 180.00, 190.00, 0.057, 2.0, "O'Neil", datetime.now()],
        ["MSFT", "Microsoft", "Tech", "BUY", 88, 85, 90, 88, 380.00, 405.00, 370.00, 395.00, 415.00, 0.066, 2.1, "Lynch", datetime.now()],
        ["GOOGL", "Google", "Tech", "HOLD", 75, 72, 78, 75, 140.00, 145.00, 136.00, 143.00, 148.00, 0.036, 1.5, "O'Neil", datetime.now()],
        ["TSLA", "Tesla", "Auto", "HOLD", 72, 70, 75, 71, 245.00, 255.00, 239.00, 250.00, 260.00, 0.041, 1.6, "Lynch", datetime.now()],
    ]

    for row in dummy_data:
        ws.append(row)

    # Create Table
    tab = Table(displayName="tbl_Signals", ref=f"A1:{get_column_letter(len(headers))}6")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Format columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.sheet_state = 'hidden'
    return ws

def create_data_prices_sheet(wb):
    """Create DATA_Prices sheet (very hidden)"""
    ws = wb.create_sheet("DATA_Prices")

    # Headers
    headers = ["ticker", "price", "change", "change_percent", "volume", "updated_at"]
    ws.append(headers)

    # Dummy data
    dummy_data = [
        ["NVDA", 485.00, 5.2, 0.0108, 45000000, datetime.now()],
        ["AAPL", 175.00, 2.3, 0.0133, 52000000, datetime.now()],
        ["MSFT", 380.00, 4.1, 0.0109, 28000000, datetime.now()],
        ["GOOGL", 140.00, 1.8, 0.0130, 31000000, datetime.now()],
        ["TSLA", 245.00, -3.2, -0.0129, 89000000, datetime.now()],
    ]

    for row in dummy_data:
        ws.append(row)

    # Create Table
    tab = Table(displayName="tbl_Prices", ref=f"A1:F6")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    # Format columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    ws.sheet_state = 'hidden'
    return ws

def create_data_context_sheet(wb):
    """Create DATA_Context sheet (very hidden)"""
    ws = wb.create_sheet("DATA_Context")

    # Headers
    headers = ["key", "value"]
    ws.append(headers)

    # Dummy data
    dummy_data = [
        ["market_regime", "bull"],
        ["regime_confidence", 0.85],
        ["breadth_sma50", 0.73],
        ["breadth_sma200", 0.65],
        ["signals_count", 247],
        ["win_rate_30d", 0.68],
        ["avg_return_30d", 0.082],
    ]

    for row in dummy_data:
        ws.append(row)

    # Create Table
    tab = Table(displayName="tbl_Context", ref="A1:B8")
    style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                          showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 15

    ws.sheet_state = 'hidden'
    return ws

def create_dashboard_sheet(wb, tier="FREE"):
    """Create Dashboard sheet"""
    ws = wb.create_sheet("Dashboard")

    # Header (A1:P3 merged)
    ws.merge_cells('A1:P3')
    header_cell = ws['A1']
    header_text = "INDICIUM SIGNALS\nSignalsSheets™ - " + tier
    header_cell.value = header_text
    header_cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
    header_cell.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Timestamp
    ws['A4'] = '=TEXT(NOW(),"DD MMM YYYY, HH:MM")&" EST"'
    ws['A4'].font = Font(color=COLORS['text_secondary'], size=10)

    # KPI Cards
    create_kpi_card(ws, "B6:D12", "MARKET REGIME",
                   '=IF(XLOOKUP("market_regime",tbl_Context[key],tbl_Context[value])="bull","🟢 BULL MARKET",IF(XLOOKUP("market_regime",tbl_Context[key],tbl_Context[value])="neutral","🟡 NEUTRAL","🔴 BEAR MARKET"))',
                   'C10', '="Breadth: "&TEXT(XLOOKUP("breadth_sma50",tbl_Context[key],tbl_Context[value]),"0%")')

    create_kpi_card(ws, "F6:H12", "SIGNALS ACTIVAS",
                   '=XLOOKUP("signals_count",tbl_Context[key],tbl_Context[value])',
                   'G10', '"en sistema"',
                   'G12', '"18 BUY hoy"')

    create_kpi_card(ws, "J6:L12", "WIN RATE",
                   '=TEXT(XLOOKUP("win_rate_30d",tbl_Context[key],tbl_Context[value]),"0%")',
                   'K10', '"(30 días)"',
                   'K12', '=TEXT(XLOOKUP("avg_return_30d",tbl_Context[key],tbl_Context[value]),"+0.0%")')

    # TOP 5 or TOP 10 Table
    top_count = 5 if tier == "FREE" else 10
    ws['A16'] = "Ticker"
    ws['B16'] = "Company"
    ws['C16'] = "Signal"
    ws['D16'] = "Score"
    ws['E16'] = "Price"
    ws['F16'] = "Target"
    ws['G16'] = "Pot.%"
    ws['H16'] = "Broker"

    # Style headers
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        cell = ws[f'{col}16']
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))

    # Data rows
    for i in range(1, top_count + 1):
        row = 17 + i
        ws[f'A{row}'] = f'=INDEX(tbl_Signals[ticker],{i})'
        ws[f'B{row}'] = f'=INDEX(tbl_Signals[company],{i})'
        ws[f'C{row}'] = f'=INDEX(tbl_Signals[signal],{i})'
        ws[f'D{row}'] = f'=INDEX(tbl_Signals[trinity_score],{i})'
        ws[f'E{row}'] = f'=INDEX(tbl_Signals[price],{i})'
        ws[f'F{row}'] = f'=INDEX(tbl_Signals[target],{i})'
        ws[f'G{row}'] = f'=INDEX(tbl_Signals[potential_return],{i})'
        ws[f'H{row}'] = "IB →"

    # Upgrade banner (only for FREE tier)
    if tier == "FREE":
        banner_start = 32

        # Set values BEFORE merging and filling
        ws[f'B{banner_start+1}'] = "💡 Desbloquea el Plan Básico ($49/mes):"
        ws[f'B{banner_start+1}'].font = Font(size=14, bold=True)

        ws[f'B{banner_start+3}'] = "✓ 25 watchlist (vs 10)"
        ws[f'B{banner_start+4}'] = "✓ 10 señales diarias (vs 5)"
        ws[f'B{banner_start+5}'] = "✓ Trade Execution Helper"
        ws[f'B{banner_start+6}'] = "✓ Portfolio Risk Manager"
        ws[f'B{banner_start+9}'] = "🔓 Upgrade Ahora →"
        ws[f'B{banner_start+9}'].font = Font(size=12, bold=True, color=COLORS['primary'])

        # Background fill (don't merge, just apply background to range)
        for row in range(banner_start, banner_start + 10):
            for col in range(1, 16):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=COLORS['upgrade_bg'], end_color=COLORS['upgrade_bg'], fill_type='solid')

    # Add Power Query note
    note_row = 45
    ws[f'A{note_row}'] = "NOTA: Power Query debe agregarse manualmente. Ver guía Day1_FINAL para código de queries."
    ws[f'A{note_row}'].font = Font(italic=True, size=9, color=COLORS['text_secondary'])

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 10

    return ws

def create_kpi_card(ws, cell_range, title, main_formula, *extra_cells_formulas):
    """Helper function to create KPI cards"""
    # Parse range
    start_cell, end_cell = cell_range.split(':')

    # Fill background
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = PatternFill(start_color=COLORS['card_bg'], end_color=COLORS['card_bg'], fill_type='solid')

    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws[cell_range]:
        for cell in row:
            cell.border = thin_border

    # Title (row 1, column 2 of range)
    start_col = ord(start_cell[0]) - ord('A') + 1
    start_row = int(start_cell[1:])

    title_cell = ws.cell(row=start_row + 1, column=start_col + 1)
    title_cell.value = title
    title_cell.font = Font(bold=True, size=11)
    title_cell.alignment = Alignment(horizontal='center')

    # Main value (row 4, column 2 of range)
    value_cell = ws.cell(row=start_row + 3, column=start_col + 1)
    value_cell.value = main_formula
    value_cell.font = Font(size=14, bold=True)
    value_cell.alignment = Alignment(horizontal='center')

    # Extra cells (if provided)
    if extra_cells_formulas:
        for i in range(0, len(extra_cells_formulas), 2):
            if i + 1 < len(extra_cells_formulas):
                cell_ref = extra_cells_formulas[i]
                formula = extra_cells_formulas[i + 1]
                ws[cell_ref] = formula
                ws[cell_ref].alignment = Alignment(horizontal='center')
                ws[cell_ref].font = Font(size=10)

def create_watchlist_sheet(wb, tier="FREE"):
    """Create Mi Watchlist sheet"""
    ws = wb.create_sheet("Mi Watchlist")

    max_rows = 10 if tier == "FREE" else 25

    # Header (A1:J3 merged)
    end_col = 'R' if tier == "BASICO" else 'J'
    ws.merge_cells(f'A1:{end_col}3')
    header_cell = ws['A1']
    header_cell.value = "⭐ MI WATCHLIST"
    header_cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
    header_cell.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Instructions
    ws['A5'] = f"Instrucciones: Ingresa tickers en columna B (máx {max_rows})"
    ws['A5'].font = Font(size=10, italic=True)
    ws.merge_cells(f'A5:{end_col}6')

    # Headers
    headers_free = ["N°", "Ticker", "Company", "Sector", "Price", "Signal", "Score", "Trend", "Notas"]
    headers_basico = headers_free + ["Entrada", "Stop", "Target", "P/L", "%Gain", "Días", "R/R", "🗑️"]

    headers = headers_basico if tier == "BASICO" else headers_free

    for idx, header in enumerate(headers):
        cell = ws.cell(row=8, column=idx + 1)
        cell.value = header
        cell.font = Font(bold=True)
        cell.border = Border(bottom=Side(style='thin'))

    # Data rows
    start_row = 10
    end_row = start_row + max_rows - 1

    for row in range(start_row, end_row + 1):
        # N°
        ws[f'A{row}'] = f'=ROW()-9'

        # Ticker (unlocked for user input)
        ws[f'B{row}'].protection = None

        # Company
        ws[f'C{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Signals[ticker],tbl_Signals[company]),"")'

        # Sector
        ws[f'D{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Signals[ticker],tbl_Signals[sector]),"")'

        # Price
        ws[f'E{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Prices[ticker],tbl_Prices[price]),"")'

        # Signal
        ws[f'F{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Signals[ticker],tbl_Signals[signal]),"")'

        # Score
        ws[f'G{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Signals[ticker],tbl_Signals[trinity_score]),"")'

        # Trend
        ws[f'H{row}'] = "↗️"

        # Notas (unlocked for user input)
        ws[f'I{row}'].protection = None

        # BASICO tier extra columns
        if tier == "BASICO":
            # Entrada (unlocked)
            ws[f'K{row}'].protection = None

            # Stop
            ws[f'L{row}'] = f'=IF(K{row}<>"",K{row}*(1-0.025),"")'

            # Target
            ws[f'M{row}'] = f'=IFERROR(XLOOKUP(B{row},tbl_Signals[ticker],tbl_Signals[target]),"")'

            # P/L (placeholder - needs quantity)
            ws[f'N{row}'] = f'=IF(AND(E{row}<>"",K{row}<>""),(E{row}-K{row})*100,"")'

            # %Gain
            ws[f'O{row}'] = f'=IF(AND(E{row}<>"",K{row}<>""),(E{row}-K{row})/K{row},"")'
            ws[f'O{row}'].number_format = '0.0%'

            # Días (placeholder)
            ws[f'P{row}'] = '""'

            # R/R
            ws[f'Q{row}'] = f'=IF(AND(M{row}<>"",K{row}<>"",L{row}<>""),(M{row}-K{row})/(K{row}-L{row}),"")'

            # Delete button
            ws[f'R{row}'] = "🗑️"

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 20

    if tier == "BASICO":
        for col in ['K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R']:
            ws.column_dimensions[col].width = 10

    # Sheet protection (unlock B and I columns)
    # Note: openpyxl doesn't fully support sheet protection, add note
    ws['A50'] = "NOTA: Protección de hoja debe configurarse manualmente en Excel (desbloquear columnas B e I)"
    ws['A50'].font = Font(italic=True, size=8, color=COLORS['text_secondary'])

    return ws

def create_placeholder_sheet(wb, sheet_name, content_note):
    """Create placeholder sheet with basic structure"""
    ws = wb.create_sheet(sheet_name)

    # Header
    ws.merge_cells('A1:P3')
    header_cell = ws['A1']
    header_cell.value = sheet_name.upper()
    header_cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFFFF')
    header_cell.fill = PatternFill(start_color=COLORS['primary'], end_color=COLORS['primary'], fill_type='solid')
    header_cell.alignment = Alignment(horizontal='center', vertical='center')

    # Content note
    ws['A6'] = content_note
    ws['A6'].font = Font(size=12, italic=True)

    return ws

def create_free_tier_excel():
    """Create SignalsSheets FREE tier Excel file"""
    print("🔨 Creando SignalsSheets_FREE.xlsm...")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create hidden sheets
    print("  → Creando hojas ocultas...")
    create_config_sheet(wb, tier="FREE")
    create_data_signals_sheet(wb)
    create_data_prices_sheet(wb)
    create_data_context_sheet(wb)

    # Create visible sheets
    print("  → Creando Dashboard...")
    create_dashboard_sheet(wb, tier="FREE")

    print("  → Creando Mi Watchlist...")
    create_watchlist_sheet(wb, tier="FREE")

    print("  → Creando hojas placeholder...")
    create_placeholder_sheet(wb, "TOP Señales",
                            "Esta hoja muestra las mejores 5 señales del día con análisis detallado.\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Índices & ETFs",
                            "Tabla con principales índices: SPY, QQQ, DIA con precios y señales.\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Metodología Trinity",
                            "Explicación educativa del Trinity Method:\n- Peter Lynch (crecimiento)\n- William O'Neil (momentum)\n- Benjamin Graham (valor)\n\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Videos Educativos",
                            "Links a videos tutoriales sobre cómo usar SignalsSheets.\nContenido pendiente de implementación.")

    # Add VBA code (stored as custom property - manual import required)
    print("  → Agregando nota sobre VBA...")
    ws_note = wb.create_sheet("_VBA_CODE_")
    ws_note['A1'] = "INSTRUCCIONES: Copiar el código VBA de abajo al módulo Module1 en Excel"
    ws_note['A1'].font = Font(bold=True, size=12, color='FFFF0000')
    ws_note['A3'] = VBA_CODE
    ws_note['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_note.column_dimensions['A'].width = 100

    # Save
    filename = "SignalsSheets_FREE.xlsm"
    wb.save(filename)
    print(f"✅ {filename} creado exitosamente!")

    return filename

def create_basico_tier_excel():
    """Create SignalsSheets BASICO tier Excel file"""
    print("\n🔨 Creando SignalsSheets_BASICO.xlsm...")

    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create hidden sheets
    print("  → Creando hojas ocultas...")
    create_config_sheet(wb, tier="BASICO")
    create_data_signals_sheet(wb)
    create_data_prices_sheet(wb)
    create_data_context_sheet(wb)

    # Create visible sheets
    print("  → Creando Dashboard...")
    create_dashboard_sheet(wb, tier="BASICO")

    print("  → Creando Mi Watchlist (25 filas)...")
    create_watchlist_sheet(wb, tier="BASICO")

    print("  → Creando hojas placeholder...")
    create_placeholder_sheet(wb, "TOP Señales",
                            "Esta hoja muestra las mejores 10 señales del día con análisis detallado.\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Índices & ETFs",
                            "Tabla con 5 índices: SPY, QQQ, DIA, IWM, VTI con precios y señales.\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Metodología Trinity",
                            "Explicación educativa del Trinity Method.\nContenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Videos Educativos",
                            "Links a videos tutoriales.\nContenido pendiente de implementación.")

    # BASICO exclusive sheets
    print("  → Creando hojas exclusivas BASICO...")
    create_placeholder_sheet(wb, "Trade Execution Helper",
                            "Herramienta para calcular tamaño de posición, stops y targets.\n" +
                            "- Dropdown ticker selector\n" +
                            "- Configuración de capital y riesgo\n" +
                            "- Sugerencias de orden calculadas\n" +
                            "Contenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Portfolio Manager",
                            "Gestión de portfolio:\n" +
                            "- Posiciones actuales\n" +
                            "- Análisis de allocation\n" +
                            "- Métricas de performance\n" +
                            "Contenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Análisis Individual",
                            "Análisis profundo de ticker individual:\n" +
                            "- Input box para ticker\n" +
                            "- Botón analizar\n" +
                            "- Output detallado\n" +
                            "Contenido pendiente de implementación.")

    create_placeholder_sheet(wb, "Market Context",
                            "Contexto de mercado:\n" +
                            "- Market Regime\n" +
                            "- Breadth Indicators\n" +
                            "- Volatility & Sentiment\n" +
                            "- Correlations\n" +
                            "Contenido pendiente de implementación.")

    # Add VBA code note
    print("  → Agregando nota sobre VBA...")
    ws_note = wb.create_sheet("_VBA_CODE_")
    ws_note['A1'] = "INSTRUCCIONES: Copiar el código VBA de abajo al módulo Module1 en Excel"
    ws_note['A1'].font = Font(bold=True, size=12, color='FFFF0000')
    ws_note['A3'] = VBA_CODE
    ws_note['A3'].alignment = Alignment(wrap_text=True, vertical='top')
    ws_note.column_dimensions['A'].width = 100

    # Save
    filename = "SignalsSheets_BASICO.xlsm"
    wb.save(filename)
    print(f"✅ {filename} creado exitosamente!")

    return filename

def main():
    """Main execution"""
    print("=" * 60)
    print("SignalsSheets Excel Generator")
    print("Trinity Method Trading Platform")
    print("=" * 60)
    print()

    try:
        # Create both files
        free_file = create_free_tier_excel()
        basico_file = create_basico_tier_excel()

        print("\n" + "=" * 60)
        print("✅ GENERACIÓN COMPLETADA")
        print("=" * 60)
        print(f"\nArchivos creados:")
        print(f"  1. {free_file}")
        print(f"  2. {basico_file}")

        print("\n📋 PASOS SIGUIENTES (MANUAL):")
        print("\n1. POWER QUERY:")
        print("   - Abrir cada archivo Excel")
        print("   - Data → Get Data → From Other Sources → Blank Query")
        print("   - Agregar queries para Cloudflare API (ver guía Day1_FINAL)")

        print("\n2. VBA CODE:")
        print("   - Alt+F11 para abrir VBA Editor")
        print("   - Ver hoja '_VBA_CODE_' en cada archivo")
        print("   - Copiar código a Module1")
        print("   - Eliminar hoja '_VBA_CODE_' después")

        print("\n3. PROTECCIÓN DE HOJAS:")
        print("   - En 'Mi Watchlist': Review → Protect Sheet")
        print("   - Desbloquear celdas B10:B19 e I10:I19 (FREE)")
        print("   - Desbloquear B10:B34, I10:I34, K10:K34 (BASICO)")

        print("\n4. OCULTAR HOJAS DATA:")
        print("   - Ejecutar macro 'MakeVeryHidden' para ocultar Config y DATA_*")

        print("\n5. HYPERLINKS (OPCIONAL):")
        print("   - Agregar links a brokers en columna 'Broker'")
        print("   - Agregar link 'Upgrade' en banner FREE")

        print("\n⚠️  LIMITACIONES DE PYTHON/OPENPYXL:")
        print("   - Power Query no puede agregarse programáticamente")
        print("   - VBA code debe importarse manualmente")
        print("   - Protección de hojas es básica (configurar en Excel)")
        print("   - Named Ranges para Config pendientes (agregar manualmente)")
        print("   - Conditional Formatting pendiente (agregar manualmente)")
        print("   - Formato XLSM creado pero VBA debe agregarse manual")

        print("\n💡 PRIORIDAD DE COMPLETADO:")
        print("   ✅ CRÍTICO: Estructura de hojas")
        print("   ✅ CRÍTICO: Tables y datos dummy")
        print("   ✅ CRÍTICO: Fórmulas Dashboard y Watchlist")
        print("   ✅ IMPORTANTE: Colores brand kit")
        print("   ⚠️  PENDIENTE: Power Query (manual)")
        print("   ⚠️  PENDIENTE: VBA import (manual)")
        print("   ⚠️  PENDIENTE: Named Ranges (manual)")
        print("   ⚠️  PENDIENTE: Conditional Formatting (manual)")

        print("\n" + "=" * 60)

    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        return 1

    return 0

if __name__ == "__main__":
    exit(main())
